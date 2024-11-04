import os
import openpyxl
import shutil
import pandas as pd
import re
import sys
import threading

from PIL.ImageOps import expand
from pdf2image import convert_from_path
from googleapiclient.discovery import build
from google.oauth2 import service_account
import tkinter as tk
from tkinter import ttk, messagebox

class PrintRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, text):
        # 將文字插入到文字框中
        self.text_widget.insert(tk.END, text)
        # 自動捲動到最新的內容
        self.text_widget.see(tk.END)

    def flush(self):
        pass  # 這裡可以留空

def get_api():
    try:
        credentials_path = './api.json'
        credentials = service_account.Credentials.from_service_account_file(
            credentials_path,
            scopes=['https://www.googleapis.com/auth/drive']
        )
        drive_service = build('drive', 'v3', credentials=credentials)
        return drive_service
    except:
        print("請確認當前資料夾中是否有api.json檔案")

def get_excel():
    try:
        excel_files = [file for file in os.listdir() if file.endswith('.xlsx')]
        excel_file_name = excel_files[0]

        if '~$' in excel_file_name:
            excel_file_name = excel_files[1]
            workbook = openpyxl.load_workbook(excel_file_name, data_only=True)
        else:
            workbook = openpyxl.load_workbook(excel_file_name, data_only=True)

        sheet = workbook["基本資料"]
        return excel_file_name , sheet
    except:
        print("請確認當前資料夾中是否有excel檔案")

def find_indices(target_headers):
    global selected_name
    indices = {}
    _, sheet = get_excel()

    for col in sheet.iter_cols(1, sheet.max_column):
        header = col[0].value
        if any(target in header for target in target_headers):
            indices[header] = col[0].column # 儲存標題的位置 (列數)

        # 以標題順序回傳找到的位置
    return [indices.get(header) for header in indices if any(target in header for target in target_headers)]

def pdf2jpg(folder_name):
    folder_path = f'./{folder_name}'
    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]

    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)

        try:
            for root_dir, _, files in os.walk(folder_path):
                for file in files:
                    if file.endswith(".pdf"):
                        pdf_path = os.path.join(root_dir, file)
                        images = convert_from_path(pdf_path)
                        for i, image in enumerate(images):
                            jpg_path = os.path.join(root_dir, f"{file[:-4]}.jpg")
                            image.save(jpg_path, 'JPEG')
                            print(f"{file}已轉換成jpg檔案")
                        os.remove(pdf_path)
        except Exception as e:
            messagebox.showerror("錯誤", f"執行發生錯誤：{e}")

def start_downlaod_multiple():
    try:
        folder_name = selected_name
        match folder_name:
            case "契約書":
                header = ["流水號", "中文姓名", "契約書"]
            case "體檢表" :
                header = ["流水號", "中文姓名", "體檢表"]
            case "安全講習":
                header = ["流水號", "中文姓名", "安全講習"]
            case _:
                pass
        num_col ,name_col ,googledriveURL_col= find_indices(header)

        # 讀取 Excel 檔案
        drive_service = get_api()
        excel_file_name , sheet = get_excel()
        pd_readexcel = pd.read_excel(excel_file_name)
        row_data = pd_readexcel.iloc[1:, (num_col) - 1]
        max_row = row_data.max()

        # 建立儲存檔案的目標資料夾
        output_folder = os.path.join('./', folder_name)
        os.makedirs(output_folder, exist_ok=True)

        for row_index, (name_cell, googledrive_cell) in enumerate(zip(
                sheet.iter_rows(min_row=2, min_col=name_col, max_col=name_col, values_only=True),
                sheet.iter_rows(min_row=2, min_col=googledriveURL_col, max_col=googledriveURL_col, values_only=True)
        )):
            if row_index == max_row:
                break

            image_urls = str(googledrive_cell).split(',')
            each_image_url = [url.strip(" '()") for url in image_urls]

            for i, image_url in enumerate(each_image_url, start=1):

                if image_url == "None":
                    print(f'跳過編號：{row_index+1}，因為googledrive_cell不包含有效的網址。')
                    continue

                if os.path.isabs(image_url):
                    name = re.sub(r"[(),' ]", '', str(name_cell))
                    file_name, file_extension = os.path.splitext(os.path.basename(image_url))
                    new_file_path = os.path.join(output_folder, f"{row_index+1}_{name}_{i}{file_extension}")
                    shutil.copy(image_url, new_file_path)
                    print(
                        f'成功將（{image_url}）複製到{folder_name}資料夾，檔案名稱：{row_index+1}_{name}_{i}{file_extension}')
                    continue

                if 'https://drive.google.com/' not in image_url:
                    continue

                file_id = image_url.split('=')[-1].strip()
                file_info = drive_service.files().get(fileId=file_id, fields="name, mimeType").execute()
                file_name = file_info["name"]
                mime_type = file_info["mimeType"]

                if mime_type == "application/pdf":
                    file_extension = ".pdf"
                elif mime_type == "image/jpeg":
                    file_extension = ".jpg"
                else:
                    file_extension = ""

                name = f"{row_index+1}_{re.sub(r'[(),\' ]', '', str(name_cell))}_{i}{file_extension}"

                response = drive_service.files().get_media(fileId=file_id)
                with open(os.path.join(output_folder, name), 'wb') as fh:
                    fh.write(response.execute())
                print(f'下載成功：{name}')

        pdf2jpg(folder_name)

        messagebox.showinfo("完成", f"所有{combo_var.get()}用{folder_name}檔案已成功下載")
        print('\n')
        download_button.config(state="normal")
    except Exception as e:
        messagebox.showerror("錯誤", f"執行發生錯誤：{e}")
        print('\n')
        download_button.config(state="normal")

def start_downlaod_single():
    try:
        folder_name = selected_name
        match folder_name:
            case "健康聲明書":
                header = ["流水號", "中文姓名", "健康聲明書"]
            case "大頭照_姓名":
                header = ["流水號", "中文姓名", "大頭照"]
            case "救生證":
                header = ["流水號", "中文姓名", "救生證照"]
            case _:
                pass
        num_col, name_col, googledriveURL_col = find_indices(header)

        # 讀取 Excel 檔案
        drive_service = get_api()
        excel_file_name, sheet = get_excel()
        pd_readexcel = pd.read_excel(excel_file_name)
        row_data = pd_readexcel.iloc[1:, (num_col) - 1]
        max_row = row_data.max()

        # 建立儲存檔案的目標資料夾
        output_folder = os.path.join('./', folder_name)
        os.makedirs(output_folder, exist_ok=True)

        for row_index, (name_cell, googledrive_cell) in enumerate(zip(
            sheet.iter_rows(min_row=2, min_col=name_col, max_col=name_col, values_only=True),
            sheet.iter_rows(min_row=2, min_col=googledriveURL_col, max_col=googledriveURL_col, values_only=True)
        )):
            if row_index == max_row:
                break

            image_url = re.sub(r"[(),' ]", '', str(googledrive_cell))

            if os.path.isabs(image_url):
                name = re.sub(r"[(),' ]", '', str(name_cell))
                file_name, file_extension = os.path.splitext(os.path.basename(image_url))
                new_file_path = os.path.join(output_folder, f"{row_index+1}_{name}{file_extension}")
                shutil.copy(image_url, new_file_path)
                print(f'成功將（{image_url}）複製到{folder_name}資料夾，檔案名稱：{row_index+1}_{name}{file_extension}')
                continue

            if 'https://drive.google.com/' not in image_url:
                print(f'跳過編號：{row_index+1}，因為googledrive_cell不包含有效的網址。')
                continue

            file_id = image_url.split('=')[-1].strip()
            file_info = drive_service.files().get(fileId=file_id, fields="name, mimeType").execute()
            file_name = file_info["name"]
            mime_type = file_info["mimeType"]

            if mime_type == "application/pdf":
                file_extension = ".pdf"
            elif mime_type == "image/jpeg":
                file_extension = ".jpg"
            else:
                file_extension = ""

            name = f"{row_index+1}_{re.sub(r'[(),\' ]', '', str(name_cell))}{file_extension}"

            response = drive_service.files().get_media(fileId=file_id)
            with open(os.path.join(output_folder, name), 'wb') as fh:
                fh.write(response.execute())
            print(f'下載成功：{name}')

        pdf2jpg(folder_name)

        messagebox.showinfo("完成", f"所有{combo_var.get()}用{folder_name}檔案已成功下載")
        print('\n')
        download_button.config(state="normal")
    except Exception as e:
        messagebox.showerror("錯誤", f"執行發生錯誤：{e}")
        print('\n')
        download_button.config(state="normal")


def start_downlaod_ID():
    try:
        folder_name = selected_name
        header = ["流水號", "身份證", "大頭照"]
        num_col, name_col, googledriveURL_col = find_indices(header)

        # 讀取 Excel 檔案
        drive_service = get_api()
        excel_file_name, sheet = get_excel()
        pd_readexcel = pd.read_excel(excel_file_name)
        row_data = pd_readexcel.iloc[1:, (num_col) - 1]
        max_row = row_data.max()

        # 建立儲存檔案的目標資料夾
        output_folder = os.path.join('./', folder_name)
        os.makedirs(output_folder, exist_ok=True)

        for row_index, (name_cell, googledrive_cell) in enumerate(zip(
            sheet.iter_rows(min_row=2, min_col=name_col, max_col=name_col, values_only=True),
            sheet.iter_rows(min_row=2, min_col=googledriveURL_col, max_col=googledriveURL_col, values_only=True)
        )):
            if row_index == max_row:
                break

            image_url = re.sub(r"[(),' ]", '', str(googledrive_cell))

            if os.path.isabs(image_url):
                name = re.sub(r"[(),' ]", '', str(name_cell))
                file_name, file_extension = os.path.splitext(os.path.basename(image_url))
                new_file_path = os.path.join(output_folder, f"{name}{file_extension}")
                shutil.copy(image_url, new_file_path)
                print(f'成功將（{image_url}）複製到{folder_name}資料夾，檔案名稱：{name}{file_extension}（編號：{row_index+1}）')
                continue

            if 'https://drive.google.com/' not in image_url:
                print(f'跳過編號：{row_index+1}，因為googledrive_cell不包含有效的網址。')
                continue

            file_id = image_url.split('=')[-1].strip()
            file_info = drive_service.files().get(fileId=file_id, fields="name, mimeType").execute()
            file_name = file_info["name"]
            mime_type = file_info["mimeType"]

            if mime_type == "application/pdf":
                file_extension = ".pdf"
            elif mime_type == "image/jpeg":
                file_extension = ".jpg"
            else:
                file_extension = ""

            name = f"{re.sub(r'[(),\' ]', '', str(name_cell))}{file_extension}"

            response = drive_service.files().get_media(fileId=file_id)
            with open(os.path.join(output_folder, name), 'wb') as fh:
                fh.write(response.execute())
            print(f'下載成功：{name}（編號：{row_index+1}）')

        pdf2jpg(folder_name)

        messagebox.showinfo("完成", f"所有{combo_var.get()}用{folder_name}檔案已成功下載")
        print('\n')
        download_button.config(state="normal")
    except Exception as e:
        messagebox.showerror("錯誤", f"執行發生錯誤：{e}")
        print('\n')
        download_button.config(state="normal")

# 創建 UI 窗口
root = tk.Tk()
root.title("Excel自動下載工具")

#大頭照_身分證只要身分證號  1_xxx
#大頭照_姓名 1_xxx
file_types = {
    "新訓": ["契約書", "體檢表", "健康聲明書", "大頭照_姓名", "大頭照_身分證"],
    "複訓": ["救生證", "安全講習", "健康聲明書", "大頭照_姓名", "大頭照_身分證"],
    "水車": ["健康聲明書", "大頭照_姓名", "大頭照_身分證"]
}

# 創建和配置UI組件
combo_var = tk.StringVar()
# 使用 Radiobutton 替代 Combobox
new_train_radio = ttk.Radiobutton(root, text="複訓", variable=combo_var, value="複訓")
retrain_radio = ttk.Radiobutton(root, text="新訓", variable=combo_var, value="新訓")
jetski_radio = ttk.Radiobutton(root, text="水車", variable=combo_var, value="水車")

new_train_radio.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
retrain_radio.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
jetski_radio.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)

tk.Label(root, text="資料夾名稱").grid(row=1, column=0, columnspan=3)
listbox = tk.Listbox(root, selectmode=tk.SINGLE)
listbox.grid(row=2, column=0, columnspan=3)

def update_listbox():
    global selected_name
    selected_option = combo_var.get()
    listbox.delete(0, tk.END)  # 清空列表框
    selected_name = None # 清除選取資料夾
    download_button.config(state="normal")

    # 根據選擇的選項更新 listbox
    if selected_option in file_types:
        items = file_types[selected_option]
        for item in items:
            listbox.insert(tk.END, item)

def listbox_binding_select(event=None):
    global selected_name
    selected_index = listbox.curselection()

    if not selected_index:
        return
    update_listbox()
    selected_name = listbox.get(selected_index)
    listbox_show = selected_name
    if "☑" not in listbox_show:
        listbox_show = selected_name + " ☑"
        # 更新 Listbox 中的項目
        listbox.delete(selected_index)
        listbox.insert(selected_index, listbox_show)
    print(f'現在選擇的是{combo_var.get()}用{selected_name}')

#設定下載模式
def set_downlaod_mode():
    try:
        download_button.config(state="disabled")
        download_thread = threading.Thread(target=set_download_files, args=(selected_name,))
        download_thread.start()
    except:
        print("請選擇資料夾名稱")
        download_button.config(state="normal")

def set_download_files(selected_name):
    match selected_name:
        case "契約書" | "體檢表" | "安全講習": #multiple files
            start_downlaod_multiple()
        case "健康聲明書" | "大頭照_姓名" | "救生證": #single file
            start_downlaod_single()
        case "大頭照_身分證": #ID card
            start_downlaod_ID()
        case _:
            print("請選擇資料夾名稱")
            download_button.config(state="normal")

new_train_radio.config(command=update_listbox)
retrain_radio.config(command=update_listbox)
jetski_radio.config(command=update_listbox)
listbox.bind("<<ListboxSelect>>", listbox_binding_select)

# 添加執行按鈕
download_button = tk.Button(root, text="下載檔案", command=set_downlaod_mode)
download_button.grid(row=3, column=0, columnspan=3)

# 右側添加文字框顯示 print 的內容
output_frame = tk.Frame(root, width=300, height=400)
output_frame.grid(row=0, column=3, rowspan=8, padx=10)

text_output = tk.Text(output_frame, wrap='word', state='normal', bg="black", fg="white")
text_output.pack(fill=tk.BOTH, expand=True)

# 將 sys.stdout 重定向到 PrintRedirector 類別
sys.stdout = PrintRedirector(text_output)

# 啟動 UI
root.mainloop()